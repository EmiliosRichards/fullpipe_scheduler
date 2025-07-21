import pandas as pd
import logging
import uuid # For RunID
from typing import Optional, List, Dict, Any, Union, Iterable # Added Iterable
from openpyxl import load_workbook # Added for smart Excel reading
import csv # Added for smart CSV reading

# Import AppConfig directly. Its __init__ handles .env loading.
# If this import fails, it's a critical setup error for the application.
from src.core.config import AppConfig
from .normalizer import apply_phone_normalization

# Configure logging
# setup_logging() might rely on environment variables loaded by AppConfig's instantiation.
try:
    from src.core.logging_config import setup_logging
    # AppConfig() is instantiated globally in config.py if needed by other modules,
    # or when an instance is created. Here, we just ensure logging is set up.
    setup_logging()
    logger = logging.getLogger(__name__)
except ImportError:
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    logger = logging.getLogger(__name__)
    logger.info("Basic logging configured for loader.py due to missing core.logging_config or its dependencies.")


def _is_row_empty(row_values: Iterable) -> bool:
    """Checks if all values in a row are None, empty string, or whitespace-only."""
    if not row_values: # Handles case where row_values itself might be None or an empty list/tuple
        return True
    return all(pd.isna(value) or (isinstance(value, str) and not value.strip()) for value in row_values)

def load_and_preprocess_data(file_path: str, app_config_instance: Optional[AppConfig] = None) -> tuple[pd.DataFrame | None, str | None, str | None]:
    """
    Loads data from a CSV or Excel file, standardizes column names based on the
    selected input profile from AppConfig, initializes new columns required for
    the pipeline, and applies initial normalization to any existing phone numbers.
    Implements smart reading for open-ended ranges to stop after a configured
    number of consecutive empty rows.

    Returns:
        A tuple containing:
        - pd.DataFrame | None: The processed DataFrame, or None on critical error.
        - str | None: The original phone column name from the selected profile.
        - str | None: The new phone column name for augmented reports, if defined.
    """
    current_config_instance: AppConfig
    if app_config_instance:
        current_config_instance = app_config_instance
    else:
        current_config_instance = AppConfig()

    original_phone_col_name_from_profile: Optional[str] = None
    skip_rows_val: Optional[int] = None
    nrows_val: Optional[int] = None
    consecutive_empty_rows_to_stop: int = 3 # Default, will be overridden by config

    if hasattr(current_config_instance, 'skip_rows_config'):
        skip_rows_val = current_config_instance.skip_rows_config
    if hasattr(current_config_instance, 'nrows_config'):
        nrows_val = current_config_instance.nrows_config
    if hasattr(current_config_instance, 'consecutive_empty_rows_to_stop'):
        consecutive_empty_rows_to_stop = current_config_instance.consecutive_empty_rows_to_stop

    log_message_parts = []
    if skip_rows_val is not None:
        log_message_parts.append(f"skipping {skip_rows_val} data rows (0-indexed)")
    if nrows_val is not None:
        log_message_parts.append(f"reading {nrows_val} data rows")
    
    smart_read_active = (nrows_val is None and consecutive_empty_rows_to_stop > 0)
    if smart_read_active:
        log_message_parts.append(f"smart read active (stop after {consecutive_empty_rows_to_stop} empty rows)")

    if log_message_parts:
        logger.info(f"Data loading configuration: {', '.join(log_message_parts)}.")
    else:
        logger.info("No specific row range configured; loading all rows (or smart read if enabled).")

    # pandas_skiprows_arg is for when not using smart read or for initial skip in some smart read scenarios
    pandas_skiprows_arg: Union[int, List[int]]
    if skip_rows_val is not None and skip_rows_val > 0:
        pandas_skiprows_arg = list(range(1, skip_rows_val + 1)) # Skips file lines 1 to skip_rows_val
    else:
        pandas_skiprows_arg = 0 # Skip no file lines after header

    df: Optional[pd.DataFrame] = None
    
    try:
        logger.info(f"Attempting to load data from: {file_path}")

        if smart_read_active:
            logger.info(f"Smart read enabled. Max consecutive empty rows: {consecutive_empty_rows_to_stop}")
            header: Optional[List[str]] = None
            data_rows: List[List[Any]] = []
            empty_row_counter = 0
            
            # Effective skip for data rows (0-indexed)
            actual_data_rows_to_skip = skip_rows_val if skip_rows_val is not None else 0

            if file_path.endswith(('.xls', '.xlsx')):
                workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
                sheet = workbook.active
                
                if sheet is None:
                    logger.warning(f"Excel file {file_path} does not have an active sheet or is empty. Returning empty DataFrame.")
                    # Ensure header is None and data_rows is empty so an empty DataFrame is created later
                    header = None
                    data_rows = []
                    # Skip the rest of the Excel processing block
                else:
                    rows_iter = sheet.iter_rows()
                    
                    # 1. Read header
                    try:
                        header_row_values = next(rows_iter)
                        header = [str(cell.value) if cell.value is not None else '' for cell in header_row_values]
                        logger.info(f"Excel header read: {header}")
                    except StopIteration:
                        logger.warning(f"Excel file {file_path} seems empty (no header row after check).")
                        # This path should ideally not be hit if sheet is not None but has no rows.
                        # If it is, header will remain None, data_rows empty.
                        pass # Allow to proceed to df creation with empty data

                    # Only proceed if header was successfully read
                    if header is not None:
                        # 2. Skip initial data rows
                        for _ in range(actual_data_rows_to_skip):
                            try:
                                next(rows_iter)
                            except StopIteration:
                                logger.info(f"Reached end of Excel file while skipping initial {actual_data_rows_to_skip} data rows.")
                                break
                        
                        # 3. Read data with empty row detection
                        for row_idx, row_values_tuple in enumerate(rows_iter):
                            current_row_values = [cell.value for cell in row_values_tuple]
                            if _is_row_empty(current_row_values):
                                empty_row_counter += 1
                                if empty_row_counter >= consecutive_empty_rows_to_stop:
                                    logger.info(f"Stopping Excel read: Found {empty_row_counter} consecutive empty rows at data row index {actual_data_rows_to_skip + row_idx}.")
                                    break
                            else:
                                empty_row_counter = 0
                                data_rows.append(current_row_values)
                
                # This block is now outside the 'else' for sheet is None,
                # but relies on 'header' and 'data_rows' being correctly set.
                if header: # If header is None (e.g. sheet was None or empty), this won't run
                    df = pd.DataFrame(data_rows, columns=header)
                    logger.info(f"Smart read from Excel resulted in {len(data_rows)} data rows.")
                elif not data_rows and header is None: # Case: sheet was None or truly empty
                    df = pd.DataFrame() # Create an empty DataFrame
                    logger.info("Smart read from Excel: sheet was None or empty, created empty DataFrame.")
                # else: # Should not happen if header is None but data_rows has content
                #    df = pd.DataFrame(data_rows) # Fallback if header is somehow None but data exists

            elif file_path.endswith('.csv'):
                with open(file_path, mode='r', encoding='utf-8', newline='') as csvfile: # Specify encoding
                    reader = csv.reader(csvfile)
                    
                    # 1. Read header
                    try:
                        header = next(reader)
                        logger.info(f"CSV header read: {header}")
                    except StopIteration:
                        logger.warning(f"CSV file {file_path} seems empty (no header row).")
                        return pd.DataFrame(), None, None
                    
                    # 2. Skip initial data rows
                    for _ in range(actual_data_rows_to_skip):
                        try:
                            next(reader)
                        except StopIteration:
                            logger.info(f"Reached end of CSV file while skipping initial {actual_data_rows_to_skip} data rows.")
                            break
                    
                    # 3. Read data with empty row detection
                    for row_idx, current_row_values in enumerate(reader):
                        if not current_row_values: # Handle completely blank lines from csv.reader
                            is_empty = True
                        else:
                            is_empty = _is_row_empty(current_row_values)

                        if is_empty:
                            empty_row_counter += 1
                            if empty_row_counter >= consecutive_empty_rows_to_stop:
                                logger.info(f"Stopping CSV read: Found {empty_row_counter} consecutive empty rows at data row index {actual_data_rows_to_skip + row_idx}.")
                                break
                        else:
                            empty_row_counter = 0
                            data_rows.append(current_row_values)
                
                if header:
                    df = pd.DataFrame(data_rows, columns=header)
                else: # Should not happen
                    df = pd.DataFrame(data_rows)
                logger.info(f"Smart read from CSV resulted in {len(data_rows)} data rows.")
            else:
                logger.error(f"Unsupported file type for smart read: {file_path}. Please use CSV or Excel.")
                return None, None, None
        else: # Original logic (fixed range or smart read disabled)
            logger.info(f"Using standard pandas read. Pandas skiprows argument: {pandas_skiprows_arg}, nrows: {nrows_val}")
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path, header=0, skiprows=pandas_skiprows_arg, nrows=nrows_val)
            elif file_path.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(file_path, header=0, skiprows=pandas_skiprows_arg, nrows=nrows_val)
            else:
                logger.error(f"Unsupported file type: {file_path}. Please use CSV or Excel.")
                return None, None, None
        
        if df is None: # Should only happen if smart read was attempted for unsupported file type
            logger.error(f"DataFrame is None after loading attempt for {file_path}. This indicates an issue with the loading logic.")
            return None, None, None

        logger.info(f"Columns loaded: {df.columns.tolist() if df is not None and not df.empty else 'N/A (DataFrame is None or empty)'}")
        
        if df.empty:
            logger.warning(f"Loaded DataFrame from {file_path} is empty. This could be due to an empty input file, all rows being skipped, or smart read stopping early.")
            # If df is empty, we still want to ensure essential columns are present for later stages if they expect them.
            # The new_columns loop later will add them.

        # --- Post-loading processing (profile-based rename_map, new_columns, etc.) ---
        active_profile_name = current_config_instance.input_file_profile_name
        profile_mappings = current_config_instance.INPUT_COLUMN_PROFILES.get(active_profile_name)

        if not profile_mappings:
            logger.error(f"Input profile '{active_profile_name}' not found in AppConfig.INPUT_COLUMN_PROFILES. Falling back to 'default'.")
            active_profile_name = "default"
            profile_mappings = current_config_instance.INPUT_COLUMN_PROFILES.get("default")
            if not profile_mappings: # Should not happen if "default" is always defined
                 logger.error("Critical: Default input profile not found. Cannot map columns.")
                 return pd.DataFrame(), None, None

        actual_rename_map = {k: v for k, v in profile_mappings.items() if not k.startswith('_') and k in df.columns}
        original_phone_col_name_from_profile = profile_mappings.get("_original_phone_column_name")
        new_phone_col_name_from_profile = profile_mappings.get("_new_phone_column_name")

        if not original_phone_col_name_from_profile and not new_phone_col_name_from_profile:
            logger.warning(f"Neither '_original_phone_column_name' nor '_new_phone_column_name' is defined for profile '{active_profile_name}'. Augmented report generation will be skipped.")
        
        if actual_rename_map:
             df.rename(columns=actual_rename_map, inplace=True)
        logger.info(f"DataFrame columns after renaming (profile: '{active_profile_name}'): {df.columns.tolist()}")

        # Ensure 'GivenPhoneNumber' column exists, as downstream processes rely on it.
        if 'GivenPhoneNumber' not in df.columns:
            logger.info("'GivenPhoneNumber' column not found in input, creating it with empty values.")
            df['GivenPhoneNumber'] = None

        new_columns = [
            "NormalizedGivenPhoneNumber", "ScrapingStatus",
            "Overall_VerificationStatus", "Original_Number_Status",
            "Primary_Number_1", "Primary_Type_1", "Primary_SourceURL_1",
            "Secondary_Number_1", "Secondary_Type_1", "Secondary_SourceURL_1",
            "Secondary_Number_2", "Secondary_Type_2", "Secondary_SourceURL_2",
            "RunID", "TargetCountryCodes"
        ]

        current_run_id = str(uuid.uuid4())

        for col in new_columns:
            if col not in df.columns:
                if col == "RunID":
                    df[col] = current_run_id
                elif col == "TargetCountryCodes":
                    # Ensure this is robust for empty df
                    df[col] = pd.Series([["DE", "AT", "CH"] for _ in range(len(df))] if not df.empty else [], dtype=object)
                elif col in ["ScrapingStatus", "Overall_VerificationStatus", "Original_Number_Status"]:
                    df[col] = "Pending"
                elif col.startswith("Primary_") or col.startswith("Secondary_"):
                    df[col] = None 
                else:
                    df[col] = None
        
        logger.info(f"Successfully loaded and structured data from {file_path}. DataFrame shape: {df.shape}")
        
        if "GivenPhoneNumber" in df.columns and not df.empty:
            df = apply_phone_normalization(df.copy(),
                                           phone_column="GivenPhoneNumber",
                                           normalized_column="NormalizedGivenPhoneNumber",
                                           region_column="TargetCountryCodes")
            logger.info("Applied initial phone number normalization.")
        elif "GivenPhoneNumber" not in df.columns:
            logger.warning("'GivenPhoneNumber' column not found. Skipping phone normalization.")
            if "NormalizedGivenPhoneNumber" not in df.columns:
                 df["NormalizedGivenPhoneNumber"] = None
        else: # df is empty
            logger.info("DataFrame is empty, skipping phone normalization.")
            if "NormalizedGivenPhoneNumber" not in df.columns: # Still ensure column exists
                 df["NormalizedGivenPhoneNumber"] = None


        return df, original_phone_col_name_from_profile, new_phone_col_name_from_profile
    except FileNotFoundError:
        logger.error(f"Error: The file {file_path} was not found.")
        return None, None, None
    except pd.errors.EmptyDataError: # This might be caught by smart read logic earlier for empty files
        logger.error(f"Error: The file {file_path} is empty (pandas EmptyDataError).")
        return pd.DataFrame(), None, None # Return empty DataFrame and Nones
    except Exception as e:
        logger.error(f"An unexpected error occurred while loading data from {file_path}: {e}", exc_info=True)
        return None, None, None