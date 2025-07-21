import os
import logging
import pandas as pd
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
from pathlib import Path
from typing import List, Dict, Optional, Any, Tuple
from collections import Counter

# Assuming AppConfig might be needed for report naming templates, will pass it if necessary.
# from src.core.config import AppConfig # AppConfig itself is not directly used, but its attributes are passed.
from src.phone_retrieval.utils.helpers import get_input_canonical_url


def write_row_attrition_report(
    run_id: str, 
    attrition_data: List[Dict[str, Any]], 
    output_dir: str, 
    canonical_domain_journey_data: Dict[str, Dict[str, Any]], 
    input_to_canonical_map: Dict[str, Optional[str]],
    logger: logging.Logger
) -> int:
    """Writes the collected row attrition data to an Excel file with auto-width columns."""
    if not attrition_data:
        logger.info("No data for row attrition report. Skipping file creation.")
        return 0

    report_filename = f"row_attrition_report_{run_id}.xlsx"
    report_path = os.path.join(output_dir, report_filename)
    
    # Create a deep copy for modification to avoid changing the original list of dicts
    attrition_data_copy = [dict(row) for row in attrition_data]

    for row_data in attrition_data_copy:
        given_url = row_data.get("GivenURL")
        row_data["Derived_Input_CanonicalURL"] = get_input_canonical_url(given_url)
        
        final_processed_domain = row_data.get("Relevant_Canonical_URLs")
        row_data["Final_Processed_Canonical_Domain"] = final_processed_domain if pd.notna(final_processed_domain) and final_processed_domain != "N/A" else None

        link_to_outcome = None
        if final_processed_domain and pd.notna(final_processed_domain) and final_processed_domain != "N/A":
            if final_processed_domain in canonical_domain_journey_data:
                link_to_outcome = final_processed_domain
            else:
                input_url_key = str(given_url) if given_url is not None else "None_GivenURL_Input"
                mapped_canonical = input_to_canonical_map.get(input_url_key)
                if mapped_canonical and mapped_canonical in canonical_domain_journey_data:
                    link_to_outcome = mapped_canonical
                else:
                    logger.debug(f"AttritionReport: Could not find domain '{final_processed_domain}' (from GivenURL: {given_url}) in canonical_domain_journey_data for linking.")
        row_data["Link_To_Canonical_Domain_Outcome"] = link_to_outcome

    report_df = pd.DataFrame(attrition_data_copy)

    columns_order = [
        "InputRowID", "CompanyName", "GivenURL",
        "Derived_Input_CanonicalURL", 
        "Final_Processed_Canonical_Domain", 
        "Link_To_Canonical_Domain_Outcome", 
        "Final_Row_Outcome_Reason", "Determined_Fault_Category",
        "Relevant_Canonical_URLs", 
        "LLM_Error_Detail_Summary",
        "Input_CompanyName_Total_Count",
        "Input_CanonicalURL_Total_Count",
        "Is_Input_CompanyName_Duplicate",
        "Is_Input_CanonicalURL_Duplicate",
        "Is_Input_Row_Considered_Duplicate",
        "Timestamp_Of_Determination"
    ]
    
    for col in columns_order:
        if col not in report_df.columns:
            report_df[col] = None 
    report_df = report_df[columns_order]

    try:
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            report_df.to_excel(writer, index=False, sheet_name='Attrition_Report')
            worksheet = writer.sheets['Attrition_Report']
            for col_idx, col_name in enumerate(report_df.columns):
                series_data = report_df.iloc[:, col_idx]
                if series_data.empty:
                    max_val_len = 0
                else:
                    lengths = series_data.astype(str).map(len)
                    max_val_len = lengths.max() if not lengths.empty else 0
                
                column_header_len = len(str(col_name))
                adjusted_width = max(max_val_len, column_header_len) + 2 
                worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
        
        logger.info(f"Row attrition report successfully saved to {report_path}")
        return len(report_df)
    except Exception as e: 
        logger.error(f"Failed to write row attrition report to {report_path}: {e}", exc_info=True)
        return 0


def write_canonical_domain_summary_report(
    run_id: str,
    domain_journey_data: Dict[str, Dict[str, Any]],
    output_dir: str,
    logger: logging.Logger
) -> int:
    """
    Writes the canonical domain journey data to an Excel file.
    """
    if not domain_journey_data:
        logger.info("No data for canonical domain summary report. Skipping file creation.")
        return 0

    report_filename = f"canonical_domain_processing_summary_{run_id}.xlsx"
    report_path = os.path.join(output_dir, report_filename)

    report_data_list = []
    for domain, data in domain_journey_data.items():
        row = {"Canonical_Domain": domain}
        # Create a copy of data to avoid modifying the original dict
        data_copy = dict(data)
        row.update(data_copy)
        report_data_list.append(row)
    
    report_df = pd.DataFrame(report_data_list)

    columns_order = [
        "Canonical_Domain", "Input_Row_IDs", "Input_CompanyNames", "Input_GivenURLs",
        "Pathful_URLs_Attempted_List", "Overall_Scraper_Status_For_Domain",
        "Total_Pages_Scraped_For_Domain", "Scraped_Pages_Details_Aggregated",
        "Regex_Candidates_Found_For_Any_Pathful", "LLM_Calls_Made_For_Domain",
        "LLM_Total_Raw_Numbers_Extracted", "LLM_Total_Consolidated_Numbers_Found",
        "LLM_Consolidated_Number_Types_Summary", "LLM_Processing_Error_Encountered_For_Domain",
        "LLM_Error_Messages_Aggregated", "Final_Domain_Outcome_Reason",
        "Primary_Fault_Category_For_Domain"
    ]

    for col in columns_order:
        if col not in report_df.columns:
            report_df[col] = None
            logger.warning(f"Column '{col}' was not found in canonical_domain_summary_report DataFrame and was initialized to None.")
    report_df = report_df[columns_order]

    for col_name in ["Input_Row_IDs", "Input_CompanyNames", "Input_GivenURLs", "Pathful_URLs_Attempted_List", "LLM_Error_Messages_Aggregated"]:
        if col_name in report_df.columns:
            report_df[col_name] = report_df[col_name].apply(lambda x: ", ".join(sorted(list(map(str, x)))) if isinstance(x, (set, list)) else x)
    
    for col_with_counter in ["Scraped_Pages_Details_Aggregated", "LLM_Consolidated_Number_Types_Summary"]:
        if col_with_counter in report_df.columns:
            report_df[col_with_counter] = report_df[col_with_counter].apply(
                lambda x: json.dumps(dict(x)) if isinstance(x, Counter) else x
            )

    try:
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            report_df.to_excel(writer, index=False, sheet_name='Canonical_Domain_Summary')
            worksheet = writer.sheets['Canonical_Domain_Summary']
            for col_idx, col_name in enumerate(report_df.columns):
                series_data = report_df[col_name]
                max_val_len = series_data.astype(str).map(len).max() if not series_data.empty else 0
                column_header_len = len(str(col_name))
                adjusted_width = max(max_val_len, column_header_len) + 5
                worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
        
        logger.info(f"Canonical domain summary report successfully saved to {report_path}")
        return len(report_df)
    except Exception as e:
        logger.error(f"Failed to write canonical domain summary report to {report_path}: {e}", exc_info=True)
        return 0


def generate_augmented_input_report(
    original_input_df: pd.DataFrame,
    processed_df: pd.DataFrame,
    output_path_augmented_excel: str,
    logger: logging.Logger,
    original_phone_col_name: Optional[str] = None,
    new_phone_col_name: Optional[str] = None
) -> None:
    """
    Generates an augmented version of the original input report.
    Updates an existing column or adds a new one with the best found phone number.
    """
    if not original_phone_col_name and not new_phone_col_name:
        logger.error("Cannot generate augmented report: No original or new phone column name provided.")
        return

    target_col = original_phone_col_name or new_phone_col_name
    logger.info(f"Starting generation of Augmented Input Report. Target column: '{target_col}'. Output: '{output_path_augmented_excel}'")

    augmented_df = original_input_df.copy()

    if original_phone_col_name and original_phone_col_name not in augmented_df.columns:
        logger.error(f"Original phone column '{original_phone_col_name}' not found in the input DataFrame. Cannot update.")
        return
    
    if new_phone_col_name and new_phone_col_name not in augmented_df.columns:
        logger.info(f"Creating new column '{new_phone_col_name}' in augmented report.")
        augmented_df[new_phone_col_name] = "" # Initialize empty column
    
    # Ensure the target column can hold strings to prevent FutureWarning
    augmented_df[target_col] = augmented_df[target_col].astype(object)

    update_map: Dict[Any, Optional[str]] = {}
    for idx, p_row in processed_df.iterrows():
        best_number: Optional[str] = None
        top_num_1 = p_row.get('Top_Number_1')
        if pd.notna(top_num_1) and isinstance(top_num_1, str) and top_num_1.strip():
            best_number = top_num_1
        update_map[idx] = best_number

    for idx, updated_phone in update_map.items():
        if idx in augmented_df.index:
            augmented_df.loc[idx, target_col] = updated_phone if updated_phone is not None else ""
        else:
            logger.warning(f"Index {idx} from processed_df not found in original_input_df for augmented report. Skipping update for this row.")
            
    try:
        with pd.ExcelWriter(output_path_augmented_excel, engine='openpyxl') as writer:
            augmented_df.to_excel(writer, index=False, sheet_name='Augmented_Input')
            worksheet = writer.sheets['Augmented_Input']
            for col_idx, column_name_header in enumerate(augmented_df.columns):
                series = augmented_df[column_name_header]
                max_len = 0
                if not series.empty:
                    str_series_lengths = series.astype(str).map(len)
                    if not str_series_lengths.empty:
                         max_len_data = str_series_lengths.max()
                         if pd.notna(max_len_data):
                            max_len = int(max_len_data)
                header_length = len(str(column_name_header))
                adjusted_width = max(max_len, header_length) + 2
                worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
        logger.info(f"Augmented Input Report successfully saved to {output_path_augmented_excel}")
    except Exception as e:
        logger.error(f"Failed to write Augmented Input Report to {output_path_augmented_excel}: {e}", exc_info=True)


def save_detailed_report(
    df_detailed_export: pd.DataFrame, 
    detailed_columns_order: List[str],
    run_id: str, 
    run_output_dir: str, 
    logger: logging.Logger
) -> int:
    """Saves the detailed LLM extractions report to an Excel file."""
    if df_detailed_export.empty:
        logger.info("No data for detailed flattened report. Skipping file creation.")
        return 0

    # Ensure all columns in detailed_columns_order exist, add if not
    for col in detailed_columns_order:
        if col not in df_detailed_export.columns:
            df_detailed_export[col] = None
    
    # Select and reorder columns
    df_to_export = df_detailed_export[detailed_columns_order].copy()

    detailed_output_filename = f"All_LLM_Extractions_Report_{run_id}.xlsx"
    detailed_output_excel_path = os.path.join(run_output_dir, detailed_output_filename)
    try:
        logger.info(f"Attempting to save detailed report to: {detailed_output_excel_path}")
        with pd.ExcelWriter(detailed_output_excel_path, engine='openpyxl') as writer:
            df_to_export.to_excel(writer, index=False, sheet_name='Detailed_Phone_Data')
            worksheet_detailed = writer.sheets['Detailed_Phone_Data']
            for col_idx, col_name in enumerate(df_to_export.columns):
                series_data = df_to_export.iloc[:, col_idx]
                if series_data.empty:
                    max_val_len = 0
                else:
                    lengths = series_data.astype(str).map(len)
                    max_val_len = lengths.max() if not lengths.empty else 0
                
                column_header_len = len(str(col_name))
                adjusted_width = max(max_val_len, column_header_len) + 2
                worksheet_detailed.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
        logger.info(f"Detailed report saved successfully to {detailed_output_excel_path}")
        return len(df_to_export)
    except Exception as e_detailed:
        logger.error(f"Error saving detailed report to {detailed_output_excel_path}: {e_detailed}", exc_info=True)
        return 0


def save_summary_report(
    df_summary_export: pd.DataFrame, 
    summary_columns_order: List[str],
    run_id: str, 
    run_output_dir: str, 
    output_excel_file_name_template: str, 
    logger: logging.Logger
) -> int:
    """Saves the summary report to an Excel file."""
    if df_summary_export.empty:
        logger.info("No data for summary report. Skipping file creation.")
        return 0
        
    # Ensure all columns in summary_columns_order exist, add if not
    for col in summary_columns_order:
        if col not in df_summary_export.columns:
            df_summary_export[col] = None # Or some other default like ""
            logger.warning(f"Summary report column '{col}' was not found in DataFrame and was initialized to None.")

    # Select and reorder columns
    df_to_export = df_summary_export[summary_columns_order].copy()


    summary_output_filename = output_excel_file_name_template.format(run_id=run_id)
    summary_output_excel_path = os.path.join(run_output_dir, summary_output_filename)
    try:
        logger.info(f"Attempting to save summary report to: {summary_output_excel_path}")
        with pd.ExcelWriter(summary_output_excel_path, engine='openpyxl') as writer:
            df_to_export.to_excel(writer, index=False, sheet_name='Phone_Validation_Summary')
            worksheet_summary = writer.sheets['Phone_Validation_Summary']
            for col_idx, col_name in enumerate(df_to_export.columns):
                series_data = df_to_export.iloc[:, col_idx]
                if series_data.empty:
                    max_val_len = 0
                else:
                    lengths = series_data.astype(str).map(len)
                    max_val_len = lengths.max() if not lengths.empty else 0
                
                column_header_len = len(str(col_name))
                adjusted_width = max(max_val_len, column_header_len) + 2
                worksheet_summary.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
        logger.info(f"Summary report saved successfully to {summary_output_excel_path}")
        return len(df_to_export)
    except Exception as e_summary:
        logger.error(f"Error saving summary report to {summary_output_excel_path}: {e_summary}", exc_info=True)
        # Optionally, re-raise or handle more gracefully if this is critical
        return 0


def save_tertiary_report(
    df_tertiary_export: pd.DataFrame, 
    tertiary_report_columns_order: List[str],
    run_id: str, 
    run_output_dir: str, 
    tertiary_report_file_name_template: str, 
    logger: logging.Logger
) -> int:
    """Saves the tertiary report (Contact Focused Report) to an Excel file."""
    if df_tertiary_export.empty:
        logger.info("No data for tertiary report. Skipping file creation.")
        return 0

    # Ensure all columns in tertiary_report_columns_order exist, add if not
    for col in tertiary_report_columns_order:
        if col not in df_tertiary_export.columns:
            df_tertiary_export[col] = None
    
    # Select and reorder columns
    df_to_export = df_tertiary_export[tertiary_report_columns_order].copy()

    tertiary_output_filename = tertiary_report_file_name_template.format(run_id=run_id)
    tertiary_output_excel_path = os.path.join(run_output_dir, tertiary_output_filename)
    try:
        logger.info(f"Attempting to save tertiary report ('Final Contacts.xlsx') to: {tertiary_output_excel_path}")
        with pd.ExcelWriter(tertiary_output_excel_path, engine='openpyxl') as writer_t:
            df_to_export.to_excel(writer_t, index=False, sheet_name='Contact_Focused_Report')
            worksheet_tertiary = writer_t.sheets['Contact_Focused_Report']
            for col_idx, col_name in enumerate(df_to_export.columns):
                series_data = df_to_export.iloc[:, col_idx]
                if series_data.empty:
                    max_val_len = 0
                else:
                    lengths = series_data.astype(str).map(len)
                    max_val_len = lengths.max() if not lengths.empty else 0
                
                column_header_len = len(str(col_name))
                adjusted_width = max(max_val_len, column_header_len) + 2
                worksheet_tertiary.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
        logger.info(f"Tertiary report saved successfully to {tertiary_output_excel_path}")
        return len(df_to_export)
    except Exception as e_tertiary:
        logger.error(f"Error saving tertiary report to {tertiary_output_excel_path}: {e_tertiary}", exc_info=True)
        return 0